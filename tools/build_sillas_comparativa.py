#!/usr/bin/env python3
"""
build_sillas_comparativa.py

Añade una hoja "Comparativa" (tabla técnica + cobertura de campos) a un
Excel ya generado por build_sillas_report.py, a partir del JSON maestro.
Solo lectura de datos ya extraídos: no vuelve a tocar Amazon ni modifica
amazon_sillas_coche.py / amazon_import.py. Las hojas existentes del Excel
(Productos, Reseñas) se conservan intactas.

Uso:
    python tools/build_sillas_comparativa.py <final.json> <final.xlsx>
"""

import json
import sys

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

ATRIBUTOS_TECNICOS = [
    ("grupo_edad", "Grupo / rango de edad"),
    ("peso_recomendado", "Peso recomendado (kg)"),
    ("isofix", "ISOFIX"),
    ("orientacion", "A contramarcha / a favor de la marcha"),
    ("reclinable", "Reclinable"),
    ("homologacion", "Homologación / normativa"),
    ("giro_360", "Giratoria 360º"),
    ("reposacabezas_regulable", "Reposacabezas regulable en altura"),
    ("tipo_arnes_proteccion", "Tipo de arnés / protección"),
    ("peso_silla", "Peso de la silla (kg)"),
    ("proteccion_lateral", "Protección lateral de impactos"),
    ("funda_lavable", "Funda lavable / desenfundable"),
    ("travel_system", "Compatibilidad capazo / travel system"),
]

NA_FILL = PatternFill(start_color="FBE2DD", end_color="FBE2DD", fill_type="solid")
NA_FONT = Font(italic=True, color="9C3B2E")
HEAD_FILL = PatternFill(start_color="1F3A5F", end_color="1F3A5F", fill_type="solid")
HEAD_FONT = Font(bold=True, color="FFFFFF")


def build_comparativa(json_path: str, xlsx_path: str) -> dict:
    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)
    productos = data["productos"]

    wb = load_workbook(xlsx_path)
    if "Comparativa" in wb.sheetnames:
        del wb["Comparativa"]
    ws = wb.create_sheet("Comparativa")

    headers = (
        ["Marca", "Modelo / producto", "Precio (€)"]
        + [label for _, label in ATRIBUTOS_TECNICOS]
        + ["Puntuación", "Nº valoraciones"]
    )
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 34

    for p in productos:
        precio = p["precio"]["actual"]
        row = [p["marca"], p["modelo"], precio if precio != "N/A" else "N/A"]
        row += [p["caracteristicas"].get(key, "N/A") for key, _ in ATRIBUTOS_TECNICOS]
        row += [p["valoraciones"]["puntuacion"], p["valoraciones"]["numero"]]
        ws.append(row)

    n_cols = len(headers)
    for r in range(2, ws.max_row + 1):
        for c in range(1, n_cols + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value == "N/A":
                cell.fill = NA_FILL
                cell.font = NA_FONT

    widths = [14, 42, 12] + [22] * len(ATRIBUTOS_TECNICOS) + [11, 15]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # --- Resumen estadístico de cobertura ---
    start_row = ws.max_row + 3
    ws.cell(row=start_row, column=1, value="Cobertura por atributo técnico").font = Font(bold=True, size=13)

    cov_header_row = start_row + 1
    cov_headers = ["Campo", "Productos con dato", "Productos con N/A", "% de cobertura"]
    for i, h in enumerate(cov_headers, start=1):
        cell = ws.cell(row=cov_header_row, column=i, value=h)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL

    n = len(productos)
    coverage = []
    for key, label in ATRIBUTOS_TECNICOS:
        con_dato = sum(1 for p in productos if p["caracteristicas"].get(key) != "N/A")
        con_na = n - con_dato
        pct = round(con_dato / n * 100) if n else 0
        coverage.append((label, con_dato, con_na, pct))

    for i, (label, con_dato, con_na, pct) in enumerate(coverage, start=cov_header_row + 1):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=con_dato)
        ws.cell(row=i, column=3, value=con_na)
        pct_cell = ws.cell(row=i, column=4, value=f"{pct}%")
        if pct >= 80:
            pct_cell.font = Font(color="1F7A4C", bold=True)
        elif pct >= 40:
            pct_cell.font = Font(color="B5691F", bold=True)
        else:
            pct_cell.font = Font(color="B23A28", bold=True)

    wb.save(xlsx_path)
    return {"n_productos": n, "coverage": coverage}


def main() -> None:
    if len(sys.argv) != 3:
        print("uso: python tools/build_sillas_comparativa.py <final.json> <final.xlsx>", file=sys.stderr)
        sys.exit(2)
    json_path, xlsx_path = sys.argv[1:3]
    result = build_comparativa(json_path, xlsx_path)
    print(f"OK: hoja 'Comparativa' añadida a {xlsx_path} ({result['n_productos']} productos)")


if __name__ == "__main__":
    main()
