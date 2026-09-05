# -*- coding: utf-8 -*-
"""Genera plan.xlsx (copia de revision editorial) a partir de plan.json."""
import json
import pathlib

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = pathlib.Path(__file__).resolve().parent
data = json.loads((BASE / "plan.json").read_text(encoding="utf-8"))

# --- Validacion ---
errors, warns = [], []
places = data["places"]
ranks = [p["rank"] for p in places]
if sorted(ranks) != list(range(1, len(places) + 1)):
    errors.append("rank duplicado, ausente o fuera de rango")
n_primary = sum(1 for p in places if p["status"] == "primary")
if n_primary != min(15, len(places)):
    errors.append("numero de primary incorrecto: %d" % n_primary)
if any(p["status"] not in ("primary", "backup") for p in places):
    errors.append("status con valor invalido")
for p in places:
    for k in ("name", "description", "rank", "status"):
        if not p.get(k):
            errors.append("campo obligatorio vacio (%s) en rank %s" % (k, p["rank"]))
    if p["evidence_count"] < p["independent_source_count"]:
        errors.append("evidence_count < independent_source_count en rank %s" % p["rank"])
names = [p["name"].strip().lower() for p in places]
if len(set(names)) != len(names):
    errors.append("lugares duplicados por nombre")
urls = [p["official_url"] for p in places if p["official_url"]]
if len(set(urls)) != len(urls):
    errors.append("lugares duplicados por official_url")
warns = [i for i in data["validation"]["issues"] if i["severity"] == "warning"]

# --- Excel ---
wb = Workbook()
ws = wb.active
ws.title = "Plan Lugo"
cols = ["rank", "status", "name", "description", "official_url", "url_verified",
        "address", "evidence_count", "independent_source_count", "sources"]
ws.append(cols)
head = Font(bold=True, color="FFFFFF")
fill = PatternFill("solid", fgColor="3C5FA0")
for c in range(1, len(cols) + 1):
    ws.cell(row=1, column=c).font = head
    ws.cell(row=1, column=c).fill = fill
for p in places:
    ws.append([
        p["rank"], p["status"], p["name"], p["description"],
        p["official_url"] or "", "si" if p["url_verified"] else "no",
        p["address"] or "", p["evidence_count"], p["independent_source_count"],
        " | ".join("%s (%s)" % (s["name"], s["domain"]) for s in p["sources"]),
    ])
widths = [6, 10, 38, 90, 46, 12, 46, 10, 12, 60]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
ws.freeze_panes = "A2"

ws2 = wb.create_sheet("Validacion")
ws2.append(["severity", "category", "message", "place_rank"])
for c in range(1, 5):
    ws2.cell(row=1, column=c).font = head
    ws2.cell(row=1, column=c).fill = fill
for i in data["validation"]["issues"]:
    ws2.append([i["severity"], i["category"], i["message"], i["place_rank"]])
for i, w in enumerate([12, 24, 120, 12], start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w
for row in ws2.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(vertical="top", wrap_text=True)

wb.save(BASE / "plan.xlsx")
print("lugares:", len(places), "| primary:", n_primary, "| backup:", len(places) - n_primary)
print("errores bloqueantes:", errors if errors else "ninguno")
print("advertencias:", len(warns))
print("xlsx ->", BASE / "plan.xlsx")
