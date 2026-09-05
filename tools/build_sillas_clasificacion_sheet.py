#!/usr/bin/env python3
"""
build_sillas_clasificacion_sheet.py

Añade (o reemplaza) una hoja "Clasificación" al Excel de auditoría ya
generado por build_sillas_report.py / build_sillas_comparativa.py, a partir
del bloque `clasificacion` (Amazon + fabricante oficial, con trazabilidad
completa) del JSON maestro. Hoja de solo lectura de datos ya extraídos: no
vuelve a tocar Amazon ni modifica amazon_sillas_coche.py / amazon_import.py.
Las hojas existentes (Productos, Reseñas, Comparativa) se conservan intactas.

Uso:
    python tools/build_sillas_clasificacion_sheet.py <final.json> <final.xlsx>
"""

import json
import sys

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

ATRIBUTOS = [
    ("isofix", "ISOFIX"),
    ("tipo_instalacion", "Tipo de instalación"),
    ("normativa", "Normativa"),
    ("grupo_r44", "Grupo"),
    ("altura_r129", "Altura (R129)"),
    ("peso", "Peso recomendado"),
    ("edad", "Edad"),
    ("orientacion", "Orientación"),
    ("giro_360", "Giro 360º"),
    ("reclinable", "Reclinable"),
    ("reposacabezas", "Reposacabezas regulable"),
    ("arnes", "Arnés / protección"),
    ("peso_silla", "Peso de la silla"),
    ("proteccion_lateral", "Protección lateral"),
    ("funda_lavable", "Funda lavable"),
]

REVISAR_FILL = PatternFill(start_color="FBE2DD", end_color="FBE2DD", fill_type="solid")
REVISAR_FONT = Font(italic=True, color="9C3B2E")
FABRICANTE_FILL = PatternFill(start_color="E2F0E4", end_color="E2F0E4", fill_type="solid")
HEAD_FILL = PatternFill(start_color="1F3A5F", end_color="1F3A5F", fill_type="solid")
HEAD_FONT = Font(bold=True, color="FFFFFF")


def build_sheet(json_path: str, xlsx_path: str) -> None:
    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)
    productos = data["productos"]

    wb = load_workbook(xlsx_path)
    if "Clasificación" in wb.sheetnames:
        del wb["Clasificación"]
    ws = wb.create_sheet("Clasificación")

    headers = ["ASIN", "Marca", "Modelo", "Identidad fabricante verificada", "Aviso nombre distinto"]
    for _, label in ATRIBUTOS:
        headers += [f"{label} — valor", f"{label} — fuente", f"{label} — revisar"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 46

    for p in productos:
        clasif = p.get("clasificacion") or {}
        fab = clasif.get("fabricante_oficial") or {}
        row = [
            p["asin"], p["marca"], p["modelo"][:60],
            "Sí" if fab.get("identidad_verificada") else ("No" if fab else ""),
            "Sí" if fab.get("aviso_nombre_diferente") else "",
        ]
        for key, _ in ATRIBUTOS:
            campo = clasif.get(key) or {}
            row += [campo.get("mostrar"), campo.get("fuente"), "Sí" if campo.get("revisar") else ""]
        ws.append(row)

        r = ws.max_row
        for i, (key, _) in enumerate(ATRIBUTOS):
            base_col = 6 + i * 3
            campo = clasif.get(key) or {}
            if campo.get("revisar"):
                for c in (base_col, base_col + 1, base_col + 2):
                    ws.cell(row=r, column=c).fill = REVISAR_FILL
                    ws.cell(row=r, column=c).font = REVISAR_FONT
            elif campo.get("fuente") and "fabricante" in campo["fuente"]:
                for c in (base_col, base_col + 1, base_col + 2):
                    ws.cell(row=r, column=c).fill = FABRICANTE_FILL

    widths = [12, 14, 40, 14, 14] + [16, 22, 9] * len(ATRIBUTOS)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "D2"

    wb.save(xlsx_path)


def main() -> None:
    if len(sys.argv) != 3:
        print("uso: python tools/build_sillas_clasificacion_sheet.py <final.json> <final.xlsx>", file=sys.stderr)
        sys.exit(2)
    json_path, xlsx_path = sys.argv[1:3]
    build_sheet(json_path, xlsx_path)
    print(f"OK: hoja 'Clasificación' añadida/actualizada en {xlsx_path}")


if __name__ == "__main__":
    main()
