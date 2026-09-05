import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

with open('tools/output/planes-familia/cantabria/plan.json', encoding='utf-8') as f:
    plan = json.load(f)

wb = Workbook()
ws = wb.active
ws.title = "Plan Cantabria"

headers = ["rank", "status", "name", "description", "official_url", "url_verified",
           "address", "evidence_count", "independent_source_count", "sources"]
ws.append(headers)
for cell in ws[1]:
    cell.font = Font(bold=True)

for place in plan["places"]:
    sources_summary = "; ".join(f"{s['name']} ({s['domain']})" for s in place["sources"])
    ws.append([
        place["rank"],
        place["status"],
        place["name"],
        place["description"],
        place["official_url"] or "",
        place["url_verified"],
        place["address"] or "",
        place["evidence_count"],
        place["independent_source_count"],
        sources_summary,
    ])

widths = [6, 9, 30, 55, 35, 11, 30, 8, 10, 45]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

for row in ws.iter_rows(min_row=2):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

wb.save('tools/output/planes-familia/cantabria/plan.xlsx')
print("saved")
