#!/usr/bin/env python3
"""
consolidar_dataset_definitivo.py

Cierra la fase de auditoria/consolidacion de la ampliacion de catalogo de
sillitas de coche (ver memoria del proyecto y CLAUDE.md). Parte de
tools/output/auditoria_30_candidatos.json (ya snapshoteado por separado
antes de ejecutar este script) y aplica TODAS las decisiones cerradas en
la sesion de consolidacion final:

  1. Saca EVOLVAFIX (Britax) de la seleccion.
  2. Anade Cybex Solution S2 i-Fix en su lugar (dato mas valorado de Cybex,
     decision explicita del usuario).
  3. Mantiene DUALFIX2 R (decision explicita: no se sustituye).
  4. Normaliza el campo "reposacabezas" a Si/No/Revisar/Sin especificar
     (booleano puro, sin numero de posiciones -- decision del usuario).
  5. Resuelve el peso de Cybex Pallas G i-Size (G2): 8,9 kg (fabricante
     oficial, confirmado manualmente; el 1,11kg de Amazon se descarta).
  6. Resuelve la normativa de Babify Onboard: R129/i-Size.
  7. Marca identidad_fabricante = "no_verificada" + fuente_datos = "amazon"
     en Chicco Kory Essential, Chicco Quasar Fix y BabyAuto Rodia Plus.

No toca amazon_import.py ni ningun archivo de la interfaz web. Genera:
  - tools/output/auditoria_30_candidatos.json (actualizado in-place)
  - tools/output/auditoria_30_candidatos_diagnostico.json (cobertura por atributo)
  - tools/output/auditoria_30_candidatos.xlsx (Comparador + Dataset completo)
  - Imprime un diff completo respecto al snapshot pasado como argumento.

Uso:
    python tools/consolidar_dataset_definitivo.py <snapshot.json> <dataset.json>
"""

import copy
import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

COMPARADOR_FIELDS = [
    ("marca", "Marca"),
    ("modelo", "Producto"),
    ("precio", "Precio (EUR)"),
    ("valoracion", "Valoracion"),
    ("n_val", "Nº valoraciones"),
    ("grupo", "Grupo"),
    ("altura", "Altura R129"),
    ("peso_recomendado", "Peso recomendado"),
    ("normativa", "Normativa"),
    ("isofix", "ISOFIX"),
    ("tipo_instalacion", "Tipo de instalacion"),
    ("orientacion", "Orientacion"),
    ("reclinable", "Reclinable"),
    ("giro_360", "Giro 360º"),
    ("reposacabezas", "Reposacabezas regulable"),
    ("arnes", "Arnes"),
    ("peso_silla", "Peso silla"),
]

INTERNO_ONLY_FIELDS = [
    ("proteccion_lateral", "Proteccion lateral"),
    ("funda_lavable", "Funda lavable"),
    ("travel_system", "Travel system"),
]

ESTADO_FIELDS = [
    "grupo", "altura", "isofix", "tipo_instalacion", "orientacion",
    "reclinable", "giro_360", "normativa", "reposacabezas", "arnes",
    "peso_silla", "proteccion_lateral", "funda_lavable", "travel_system",
]

REPOSACABEZAS_SI = {
    "B0BZ89X9YC",  # I-GROW
    "B0FGDM2ZTS",  # SAFETY FIX 3 PRO
    "B0CZ47FKZL",  # Tanza
    "B0CWS19269",  # Emerald 360 S
    "B0F3P61P8C",  # RevolveFix Plus 360
    "B0GGZYY3BZ",  # Marvel RoadSafe
    "B0C58QGJNG",  # Solution X i-Fix
    "B0BMGN89G3",  # SlimFit R129
    "B0CYQD2H66",  # Junior Maxi i-Size
    "B0CJCC2YYP",  # Bastiaan i-Size
    "B0CHS67WJ3",  # HUGO i-Size
    "B0D5QSYXPH",  # LEVI ONE i-Size
    "B0FL7MMYHN",  # i-MOOVE 2
    "B0DYDY6C5Y",  # i-PASS
    "B0DS6C4XR8",  # i-FLIT
    "B0FLDLZTHJ",  # Belem
    "B0FLDT4541",  # Bogota
    "B07RYWKS9Z",  # Onboard (Babify)
    # Resueltos hoy en la auditoria de reposacabezas:
    "B0C6R47BCV",  # Pearl 360
    "B0DB2J1HMK",  # Cloud G i-Size Comfort
    "B0DPHRJ1YH",  # Pallas G i-Size (G2)
    "B07QLSYS2Y",  # DUALFIX2 R
    "B0CP2QYQK8",  # BABY-SAFE CORE
    "B0C6KZCY6Z",  # Unico EVO I'Size Classic
    "B0CQP3VKSK",  # Rodia Plus i-Size
    "B0B97Z1C6M",  # Solution S2 i-Fix (nuevo, sustituye a EVOLVAFIX)
}
REPOSACABEZAS_NO = {"B09FQ3PYZX"}  # Alzador Coche Grupo 2/3 (Jovikids)
REPOSACABEZAS_SIN_ESPECIFICAR = {
    "B0F1R26D8L",  # I-BOOST 2
    "B0C6KXB9LL",  # Kory i-Size Essential
    "B0DW9GD5CG",  # Quasar Fix i-Size
}

# Reclinable: mismo tratamiento que reposacabezas -- booleano puro (Si/No/
# Sin especificar), sin numero de posiciones. Estos 14 ya tenian evidencia
# positiva (con detalle de posiciones); se normaliza el texto, sin volver
# a investigar nada.
RECLINABLE_SI = {
    "B0BZ89X9YC",  # I-GROW
    "B0CWS19269",  # Emerald 360 S
    "B0F3P61P8C",  # RevolveFix Plus 360
    "B0C58QGJNG",  # Solution X i-Fix
    "B0C6R47BCV",  # Pearl 360
    "B0DB2J1HMK",  # Cloud G i-Size Comfort
    "B0DPHRJ1YH",  # Pallas G i-Size (G2)
    "B07QLSYS2Y",  # DUALFIX2 R
    "B0C6KZCY6Z",  # Unico EVO I'Size Classic
    "B0CQP3VKSK",  # Rodia Plus i-Size
    "B0BMGN89G3",  # SlimFit R129
    "B0CJCC2YYP",  # Bastiaan i-Size
    "B0FL7MMYHN",  # i-MOOVE 2
    "B07RYWKS9Z",  # Onboard (Babify)
    # Resueltos en la auditoria de reclinable (13-ago-2026):
    "B0DS6C4XR8",  # i-FLIT (KikkaBoo) -- 2 posiciones confirmadas en Amazon
    "B0B97Z1C6M",  # Solution S2 i-Fix (Cybex) -- "Respaldo regulable" en Amazon
    "B0FGDM2ZTS",  # SAFETY FIX 3 PRO (Kinderkraft) -- manual oficial confirma
}
RECLINABLE_NO = {
    "B09FQ3PYZX",  # Alzador Coche Grupo 2/3 (Jovikids) -- tabla comparativa Amazon
}
# Maxi-Cosi Tanza (B0CZ47FKZL): decision explicita del usuario (13-ago-2026)
# de dejarlo en "Sin especificar" pese a una resena de cliente que sugeria
# "No" -- una resena no es fuente oficial (Amazon/fabricante) valida.
# El resto (11, "Sin especificar") se deja tal cual -- sin evidencia
# fiable en ninguna fuente admitida por el proyecto.

IDENTIDAD_NO_VERIFICADA = {"B0C6KXB9LL", "B0DW9GD5CG", "B0CQP3VKSK"}

# Auditoria de "arnes" (13-ago-2026): resuelve los 5 "Sin especificar"
# restantes (DUALFIX2 R se deja intacto, no se reinvestiga). HUGO pasa a
# ser una NUEVA contradiccion: el campo estructurado "Tipo de cinturon de
# seguridad" de Amazon dice "5 puntos", pero los bullets, la descripcion y
# la web oficial de Lionelo coinciden en que es solo cinturon de 3 puntos
# del coche -- no se resuelve automaticamente aunque el peso de la
# evidencia apunte a un lado, siguiendo la regla del proyecto.
ARNES_UPDATES = {
    "B0F1R26D8L": ("Cinturón del vehículo (3 puntos)", None),  # I-BOOST 2
    "B0DS6C4XR8": ("Arnés de 5 puntos", None),  # i-FLIT
    "B0CJCC2YYP": ("Arnés de 5 puntos", None),  # Bastiaan i-Size
    "B0D5QSYXPH": (
        "Arnés de 5 puntos (76-105cm) / cinturón de 3 puntos (100-150cm)",
        None,
    ),  # LEVI ONE i-Size
    "B0CHS67WJ3": (
        "Cinturón del vehículo (3 puntos)",
        (
            "✅ RESUELTO manualmente (13-ago-2026): el campo estructurado "
            "\"Tipo de cinturón de seguridad\" de la ficha técnica de Amazon "
            "decía \"5 puntos\", pero el resto del contenido visible de la "
            "misma ficha (bullets, descripción: \"Montaje: Montaje con "
            "cinturones de seguridad de 3 puntos\") y la web oficial de "
            "Lionelo (es.lionelo.com) coincidian en \"cinturón de 3 puntos "
            "del coche\", sin mencionar arnés en ningún punto. El usuario "
            "confirmó manualmente en el chat que en el rango 100-150cm el "
            "montaje es con ISOFIX+cinturón de 3 puntos o solo cinturón de "
            "3 puntos (coherente con tipo_instalacion ya registrado), lo "
            "que corrobora que la sujeción del niño es el cinturón de 3 "
            "puntos del coche, no un arnés propio de la silla. Se descarta "
            "el campo estructurado de Amazon (\"5 puntos\") como erróneo -- "
            "ese mismo campo parece poco fiable en sillas grupo 2/3 o "
            "multietapa (mismo patrón detectado en LEVI ONE)."
        ),
    ),  # HUGO i-Size
    "B0FLDT4541": (
        "Cinturón del vehículo",
        (
            "✅ CORREGIDO (14-ago-2026): el valor anterior (\"ISOFIX o "
            "cinturón del vehículo\") era un error de copia -- identico al "
            "campo tipo_instalacion de este mismo producto, y ISOFIX nunca "
            "puede ser un tipo de arnés (ancla la silla, no sujeta al "
            "niño). Investigado en Amazon España (única fuente disponible: "
            "Nania/Team Tex cerró en 2024, nania.fr/nania.com son dominios "
            "en venta sin contenido del fabricante): título, bullets y "
            "descripción solo mencionan ISOFIX/cinturón para el anclaje de "
            "la silla; es un alzador (booster) convertible a \"elevador "
            "sin respaldo a partir de 135cm\", categoría que por diseño no "
            "lleva arnés de 5 puntos. Coherente con su hermano de línea "
            "Nania Belem (mismo grupo 2/3, sin ISOFIX), que ya tenía "
            "correctamente \"Cinturón del vehículo\"."
        ),
    ),  # Bogota (Nania)
}

EVOLVAFIX_ASIN = "B0C3HJNK2M"

SOLUTION_S2_IFIX = {
    "asin": "B0B97Z1C6M",
    "marca": "Cybex",
    "modelo": "Solution S2 i-Fix",
    "actual_lote1": False,
    "precio": 149.95,
    "valoracion": 4.8,
    "n_val": 4098,
    "grupo": "Sin especificar",
    "edad": "3-12 años aprox.",
    "altura": "100-150 cm",
    "peso_recomendado": "15-50 kg",
    "isofix": "Sí (opcional, también cinturón)",
    "tipo_instalacion": "Cinturón + ISOFIX opcional",
    "orientacion": "A favor de la marcha",
    "reclinable": "Sin especificar",
    "giro_360": "Sin especificar",
    "normativa": "R129 / i-Size",
    "reposacabezas": "Sí",
    "arnes": "Arnés de 3 puntos",
    "peso_silla": "6,2 kg",
    "proteccion_lateral": "Sí (Sistema L.S.P., confirmado en Amazon)",
    "funda_lavable": "Sí (lavable a máquina 30°C, confirmado en Amazon)",
    "travel_system": "Sin especificar",
    "estado_auditoria": (
        "✅ Sustituye a EVOLVAFIX por decisión del usuario (13-ago-2026): "
        "es el Cybex con mas valoraciones en Amazon.es (4.098, verificado en "
        "vivo), elegido explicitamente por ese criterio tras descartar "
        "EVOLVAFIX por disponibilidad (3-7 meses de espera). Nota de "
        "redundancia: mismo rango 100-150cm que Solution X i-Fix, ya en la "
        "seleccion -- aceptado conscientemente por el usuario. Normativa "
        "R129/03 confirmada solo por fabricante oficial (cybex-online.com, "
        "manual de usuario), Amazon no la menciona explicitamente."
    ),
}


def load(path: str) -> dict:
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def save(data: dict, path: str) -> None:
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write("\n")


def normalizar_reposacabezas(productos: list) -> None:
    for p in productos:
        asin = p["asin"]
        if asin in REPOSACABEZAS_SI:
            p["reposacabezas"] = "Sí"
        elif asin in REPOSACABEZAS_NO:
            p["reposacabezas"] = "No"
        elif asin in REPOSACABEZAS_SIN_ESPECIFICAR:
            p["reposacabezas"] = "Sin especificar"
        else:
            raise ValueError(f"ASIN sin clasificar en reposacabezas: {asin} ({p['modelo']})")


def normalizar_reclinable(productos: list) -> None:
    for p in productos:
        if p["asin"] in RECLINABLE_SI:
            p["reclinable"] = "Sí"
        elif p["asin"] in RECLINABLE_NO:
            p["reclinable"] = "No"


def aplicar_pallas_g2(productos: list) -> None:
    for p in productos:
        if p["asin"] == "B0DPHRJ1YH":
            p["peso_silla"] = "8,9 kg"
            p["peso_silla_fuente"] = "fabricante_oficial (confirmado manualmente)"
            p["estado_auditoria"] = (
                "✅ RESUELTO manualmente (13-ago-2026): peso real 8,9 kg, "
                "confirmado en dos paginas independientes de cybex-online.com "
                "(cs-go-pallas-g2.html y CS_GO_Pallas_G_i-Size_EN.html), "
                "coincidentes entre si. El '1,11 Kilogramos' que muestra "
                "Amazon (repetido en dos bloques de la misma ficha) se "
                "descarta por fisicamente inverosimil para esta silla -- "
                "no se usa como dato valido."
            )
            return
    raise ValueError("No se encontro Pallas G i-Size (G2) en el dataset")


def aplicar_dualfix2r_decision_final(productos: list) -> None:
    for p in productos:
        if p["asin"] == "B07QLSYS2Y":
            p["estado_auditoria"] += (
                " | DECISION FINAL (13-ago-2026): se investigo sustituir por "
                "DUALFIX PLUS (modelo mas actual segun britax-romer.es, "
                "premio STIWA-ADAC 06/2023) pero DUALFIX PLUS tiene la misma "
                "espera de 3-7 meses en Amazon.es que llevo a descartar "
                "EVOLVAFIX; DUALFIX PRO esta directamente 'No disponible'; "
                "DUALFIX 2 Z-LINE esta disponible pero con solo 42 "
                "valoraciones y 3 unidades en stock. El usuario confirmo "
                "explicitamente mantener DUALFIX2 R sin sustitucion: sigue "
                "disponible en Amazon.es ahora (4,7★/1.809 valoraciones, "
                "solo 4 en stock) pese a estar descatalogado por el "
                "fabricante (confirmado por OCU)."
            )
            return
    raise ValueError("No se encontro DUALFIX2 R en el dataset")


def aplicar_babify_onboard(productos: list) -> None:
    for p in productos:
        if p["asin"] == "B07RYWKS9Z":
            p["normativa"] = "R129 / i-Size"
            p["estado_auditoria"] = (
                "✅ RESUELTO (13-ago-2026): normativa R129/i-Size confirmada "
                "por el titulo del listing de Amazon para este ASIN exacto "
                "(\"i-Size ECE R129\"). La mencion a R44/04 detectada en la "
                "auditoria previa procedia de manuales de modelos HERMANOS "
                "de la marca (\"Onboard Swivel 360\", \"Isofix Reclining Car "
                "Seat\" -- productos distintos, clasificados por peso, no "
                "por altura), no del manual de este producto \"Onboard\" "
                "40-150cm. Sin web de fabricante oficial independiente "
                "localizable (marca de marketplace); no se encontro numero "
                "de homologacion E-mark en ninguna fuente accesible."
            )
            return
    raise ValueError("No se encontro Babify Onboard en el dataset")


def aplicar_identidad_no_verificada(productos: list) -> None:
    notas = {
        "B0C6KXB9LL": "Kory i-Size Essential",
        "B0DW9GD5CG": "Quasar Fix i-Size",
        "B0CQP3VKSK": "Rodia Plus i-Size",
    }
    for p in productos:
        if p["asin"] in IDENTIDAD_NO_VERIFICADA:
            p["identidad_fabricante"] = "no_verificada"
            p["fuente_datos"] = "amazon"
            if "no se sigue buscando" not in p["estado_auditoria"]:
                p["estado_auditoria"] += (
                    " | Decision de cierre (13-ago-2026): identidad exacta "
                    "con el fabricante no verificada -- no se sigue buscando "
                    "indefinidamente ni se usan specs de un modelo parecido "
                    "del fabricante. El producto se mantiene en el catalogo "
                    "usando exclusivamente datos de Amazon."
                )
    faltan = IDENTIDAD_NO_VERIFICADA - {p["asin"] for p in productos if p["asin"] in notas}
    if faltan:
        raise ValueError(f"ASIN de identidad no verificada no encontrados: {faltan}")


def aplicar_arnes_updates(productos: list) -> None:
    aplicados = set()
    for p in productos:
        if p["asin"] in ARNES_UPDATES:
            nuevo_valor, nota = ARNES_UPDATES[p["asin"]]
            p["arnes"] = nuevo_valor
            if nota:
                p["estado_auditoria"] += " | " + nota
            aplicados.add(p["asin"])
    faltan = set(ARNES_UPDATES) - aplicados
    if faltan:
        raise ValueError(f"ASIN de arnes no encontrados: {faltan}")


def sacar_evolvafix_meter_solution_s2(productos: list) -> list:
    idx_evolvafix = next((i for i, p in enumerate(productos) if p["asin"] == EVOLVAFIX_ASIN), None)
    if idx_evolvafix is None:
        raise ValueError("EVOLVAFIX no encontrado -- no se puede confirmar que se elimina")
    productos.pop(idx_evolvafix)

    # Insertar justo despues de Solution X i-Fix para mantener los Cybex agrupados
    idx_solution_x = next(i for i, p in enumerate(productos) if p["asin"] == "B0C58QGJNG")
    productos.insert(idx_solution_x + 1, copy.deepcopy(SOLUTION_S2_IFIX))
    return productos


def validar(productos: list) -> None:
    assert len(productos) == 30, f"Se esperaban 30 productos, hay {len(productos)}"
    asins = [p["asin"] for p in productos]
    assert len(asins) == len(set(asins)), "ASIN duplicado en el dataset"
    assert EVOLVAFIX_ASIN not in asins, "EVOLVAFIX no deberia estar en el dataset final"

    por_marca = {}
    for p in productos:
        por_marca.setdefault(p["marca"], []).append(p["modelo"])

    britax = por_marca.get("Britax Römer", [])
    assert set(britax) == {"DUALFIX2 R", "BABY-SAFE CORE"}, f"Britax inesperado: {britax}"

    cybex = por_marca.get("Cybex", [])
    assert set(cybex) == {
        "Solution X i-Fix", "Solution S2 i-Fix", "Cloud G i-Size Comfort", "Pallas G i-Size (G2)"
    }, f"Cybex inesperado: {cybex}"

    limite_3 = {
        "Kinderkraft", "Maxi-Cosi", "Bébé Confort", "Britax Römer",
        "Chicco", "BabyAuto", "Graco", "Lionelo", "KikkaBoo", "Jovikids",
    }
    limite_2 = {"Nania"}
    limite_1 = {"Babify"}
    # Excepcion explicita aprobada por el usuario (13-ago-2026): Cybex pasa
    # a 4 al absorber el hueco que deja EVOLVAFIX (Britax baja a 2), en vez
    # de excluir un cuarto Cybex solo para respetar el limite general.
    assert len(cybex) == 4, f"Cybex deberia tener 4 (excepcion aprobada): {cybex}"
    for marca, modelos in por_marca.items():
        if marca in limite_3:
            assert len(modelos) <= 3, f"{marca} excede el limite de 3: {modelos}"
        elif marca in limite_2:
            assert len(modelos) <= 2, f"{marca} excede el limite de 2: {modelos}"
        elif marca in limite_1:
            assert len(modelos) <= 1, f"{marca} excede el limite de 1: {modelos}"

    return por_marca


def diff_datasets(antes: dict, despues: dict) -> dict:
    antes_por_asin = {p["asin"]: p for p in antes["productos"]}
    despues_por_asin = {p["asin"]: p for p in despues["productos"]}

    eliminados = sorted(set(antes_por_asin) - set(despues_por_asin))
    anadidos = sorted(set(despues_por_asin) - set(antes_por_asin))
    comunes = set(antes_por_asin) & set(despues_por_asin)

    celdas_modificadas = []
    for asin in sorted(comunes):
        a, d = antes_por_asin[asin], despues_por_asin[asin]
        campos = set(a.keys()) | set(d.keys())
        for campo in sorted(campos):
            va, vd = a.get(campo), d.get(campo)
            if va != vd:
                celdas_modificadas.append({
                    "asin": asin,
                    "modelo": d.get("modelo", a.get("modelo")),
                    "campo": campo,
                    "antes": va,
                    "despues": vd,
                })

    return {
        "eliminados": [{"asin": a, "modelo": antes_por_asin[a]["modelo"]} for a in eliminados],
        "anadidos": [{"asin": a, "modelo": despues_por_asin[a]["modelo"]} for a in anadidos],
        "productos_con_cambios": sorted({c["asin"] for c in celdas_modificadas}),
        "atributos_afectados": sorted({c["campo"] for c in celdas_modificadas}),
        "num_celdas_modificadas": len(celdas_modificadas),
        "celdas": celdas_modificadas,
    }


def build_diagnostico(productos: list) -> dict:
    diag = {"total_productos": len(productos), "cobertura": {}}
    for campo in ESTADO_FIELDS:
        si = no = revisar = sin_esp = 0
        for p in productos:
            v = str(p.get(campo, "Sin especificar"))
            vl = v.strip().lower()
            if vl.startswith("revisar") or "revisar" in vl and "confirmado" not in vl and "resuelto" not in vl:
                revisar += 1
            elif vl.startswith("sin especificar") or vl == "":
                sin_esp += 1
            elif vl.startswith("no") and campo in ("isofix", "reposacabezas", "giro_360", "reclinable"):
                no += 1
            else:
                si += 1
        diag["cobertura"][campo] = {
            "si_valor_concreto": si,
            "no": no,
            "revisar": revisar,
            "sin_especificar": sin_esp,
        }
    return diag


def write_excel(productos: list, out_path: str) -> None:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Comparador"
    headers = [label for _, label in COMPARADOR_FIELDS]
    ws1.append(headers)
    for c in range(1, len(headers) + 1):
        ws1.cell(row=1, column=c).font = Font(bold=True)
    for p in productos:
        ws1.append([p.get(key, "") for key, _ in COMPARADOR_FIELDS])
    for i, (_, label) in enumerate(COMPARADOR_FIELDS, start=1):
        ws1.column_dimensions[get_column_letter(i)].width = max(14, len(label) + 2)

    ws2 = wb.create_sheet("Dataset completo (interno)")
    all_fields = [("asin", "ASIN"), ("marca", "Marca"), ("modelo", "Producto")]
    all_fields += [(k, v) for k, v in COMPARADOR_FIELDS if k not in ("marca", "modelo")]
    all_fields += INTERNO_ONLY_FIELDS
    all_fields += [
        ("actual_lote1", "Ya en Lote 1"),
        ("identidad_fabricante", "Identidad fabricante"),
        ("fuente_datos", "Fuente datos"),
        ("peso_silla_fuente", "Fuente peso silla"),
        ("estado_auditoria", "Estado auditoria"),
    ]
    ws2.append([label for _, label in all_fields])
    for c in range(1, len(all_fields) + 1):
        ws2.cell(row=1, column=c).font = Font(bold=True)
    for p in productos:
        ws2.append([p.get(key, "") for key, _ in all_fields])
    for i, (_, label) in enumerate(all_fields, start=1):
        width = 60 if label == "Estado auditoria" else max(12, len(label) + 2)
        ws2.column_dimensions[get_column_letter(i)].width = width

    wb.save(out_path)


def main():
    if len(sys.argv) != 3:
        print("Uso: python consolidar_dataset_definitivo.py <snapshot.json> <dataset.json>", file=sys.stderr)
        sys.exit(1)

    snapshot_path, dataset_path = sys.argv[1], sys.argv[2]
    antes = load(snapshot_path)
    data = load(dataset_path)
    productos = data["productos"]

    productos = sacar_evolvafix_meter_solution_s2(productos)
    normalizar_reposacabezas(productos)
    normalizar_reclinable(productos)
    aplicar_pallas_g2(productos)
    aplicar_dualfix2r_decision_final(productos)
    aplicar_babify_onboard(productos)
    aplicar_identidad_no_verificada(productos)
    aplicar_arnes_updates(productos)

    por_marca = validar(productos)

    data["productos"] = productos
    data["_nota"] = (
        data["_nota"]
        + " [CONSOLIDACION FINAL 13-ago-2026: EVOLVAFIX sustituido por Cybex "
        "Solution S2 i-Fix; DUALFIX2 R se mantiene (decision explicita, no "
        "sustituido por DUALFIX PLUS); reposacabezas normalizado a "
        "Si/No/Sin especificar (booleano, sin numero de posiciones); peso "
        "Pallas G2 y normativa Babify Onboard resueltos; identidad de "
        "fabricante marcada explicitamente como no verificada en Kory "
        "Essential, Quasar Fix y Rodia Plus. Este dataset es la fuente de "
        "verdad para el comparador -- ver informe de consolidacion.]"
    )

    save(data, dataset_path)

    diagnostico_path = str(Path(dataset_path).with_name("auditoria_30_candidatos_diagnostico.json"))
    diagnostico = build_diagnostico(productos)
    diagnostico["productos_por_marca"] = {k: sorted(v) for k, v in por_marca.items()}
    save(diagnostico, diagnostico_path)

    excel_path = str(Path(dataset_path).with_suffix(".xlsx"))
    write_excel(productos, excel_path)

    diff = diff_datasets(antes, data)
    diff_path = str(Path(dataset_path).with_name("auditoria_30_candidatos_diff.json"))
    save(diff, diff_path)

    print(json.dumps({
        "total_productos": len(productos),
        "por_marca": {k: len(v) for k, v in por_marca.items()},
        "eliminados": diff["eliminados"],
        "anadidos": diff["anadidos"],
        "num_celdas_modificadas": diff["num_celdas_modificadas"],
        "atributos_afectados": diff["atributos_afectados"],
        "productos_con_cambios": len(diff["productos_con_cambios"]),
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
