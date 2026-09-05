#!/usr/bin/env python3
"""
build_sillas_report.py

Post-procesado del JSON crudo que produce amazon_sillas_coche.py: fusiona
los resumenes editoriales (redactados por Claude Code a partir de
opiniones.muestra, ver CLAUDE.md/memoria del proyecto) y genera el JSON
maestro final y un Excel de revision con dos hojas (Productos, Resenas).

No forma parte del scraper ni lo modifica: es una herramienta separada de
la capa editorial, tal como se acordo (Capa 1 extraccion / Capa 2 editorial
/ informe).

Uso:
    python tools/build_sillas_report.py <raw.json> <resumenes.json> <salida_base>

    <raw.json>        Salida de amazon_sillas_coche.py (--out ...)
    <resumenes.json>  {"ASIN": "resumen editorial...", ...}
    <salida_base>     Prefijo de los archivos de salida (sin extension):
                       genera <salida_base>.json y <salida_base>.xlsx
"""

import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

CARACTERISTICAS_ORDEN = [
    ("grupo_edad", "Grupo / rango edad"),
    ("peso_recomendado", "Peso recomendado"),
    ("isofix", "ISOFIX"),
    ("orientacion", "Orientación"),
    ("reclinable", "Reclinable"),
    ("homologacion", "Homologación"),
    ("giro_360", "Giro 360º"),
    ("reposacabezas_regulable", "Reposacabezas regulable"),
    ("tipo_arnes_proteccion", "Tipo arnés/protección"),
    ("peso_silla", "Peso silla"),
    ("proteccion_lateral", "Protección lateral"),
    ("funda_lavable", "Funda lavable"),
    ("travel_system", "Travel system/capazo"),
]


def build_master_json(raw_path: str, resumenes_path: str) -> dict:
    with open(raw_path, encoding="utf-8") as f:
        raw = json.load(f)
    with open(resumenes_path, encoding="utf-8") as f:
        resumenes = json.load(f)

    faltan = []
    for p in raw["productos"]:
        resumen = resumenes.get(p["asin"])
        if resumen is None:
            faltan.append(p["asin"])
        p["opiniones"]["resumen"] = resumen

    if faltan:
        print(f"[AVISO] sin resumen editorial: {faltan}", file=sys.stderr)

    return raw


def write_master_json(data: dict, out_path: str) -> None:
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def write_excel(data: dict, out_path: str) -> None:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Productos"
    headers = (
        ["ASIN", "Marca", "Modelo/nombre", "URL Amazon", "Precio actual (€)",
         "Precio anterior (€)", "Puntuación", "Nº valoraciones"]
        + [label for _, label in CARACTERISTICAS_ORDEN]
        + ["Estado", "Nº reseñas extraídas", "Resumen opiniones (editorial)"]
    )
    ws1.append(headers)
    for cell in ws1[1]:
        cell.font = Font(bold=True)

    for p in data["productos"]:
        row = [
            p["asin"], p["marca"], p["modelo"], p["url"],
            p["precio"]["actual"], p["precio"]["anterior"],
            p["valoraciones"]["puntuacion"], p["valoraciones"]["numero"],
        ]
        row += [p["caracteristicas"].get(key, "N/A") for key, _ in CARACTERISTICAS_ORDEN]
        row += [p["estado"], len(p["opiniones"]["muestra"]), p["opiniones"]["resumen"]]
        ws1.append(row)
        ws1.cell(row=ws1.max_row, column=len(headers)).alignment = Alignment(wrap_text=True, vertical="top")

    widths = [12, 14, 40, 45, 14, 16, 11, 14] + [16] * len(CARACTERISTICAS_ORDEN) + [18, 12, 60]
    for i, w in enumerate(widths, start=1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = "A2"

    ws2 = wb.create_sheet("Reseñas")
    ws2.append(["ASIN", "Marca", "Modelo/nombre", "Nº reseña", "Texto de la reseña"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    for p in data["productos"]:
        for i, review in enumerate(p["opiniones"]["muestra"], start=1):
            ws2.append([p["asin"], p["marca"], p["modelo"], i, review])
            ws2.cell(row=ws2.max_row, column=5).alignment = Alignment(wrap_text=True, vertical="top")

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 40
    ws2.column_dimensions["D"].width = 10
    ws2.column_dimensions["E"].width = 90
    ws2.freeze_panes = "A2"

    wb.save(out_path)


def main() -> None:
    if len(sys.argv) != 4:
        print("uso: python tools/build_sillas_report.py <raw.json> <resumenes.json> <salida_base>", file=sys.stderr)
        sys.exit(2)

    raw_path, resumenes_path, salida_base = sys.argv[1:4]

    data = build_master_json(raw_path, resumenes_path)
    write_master_json(data, f"{salida_base}.json")
    write_excel(data, f"{salida_base}.xlsx")
    print(f"OK: {salida_base}.json / {salida_base}.xlsx ({len(data['productos'])} productos)")


if __name__ == "__main__":
    main()
